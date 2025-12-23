"""Flask routes for web interface - Updated for multi-vendor multi-switch support"""
from flask import Blueprint, render_template, request, jsonify, send_file
from app.ssh_manager import SSHManager
from app.config_parser import ConfigParser
from app.interface_translator import InterfaceTranslator
from app.template_generator import TemplateGenerator
from app.universal_interface_parser import UniversalInterfaceParser
import logging
import io
import os
import re
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

bp = Blueprint('main', __name__)


@bp.route('/')
def index():
    """Main web interface"""
    return render_template('index.html')


@bp.route('/extract_config', methods=['POST'])
def extract_config():
    """Extract configuration from old switch via SSH"""
    try:
        data = request.get_json()

        # Validate input
        required_fields = ['ip', 'username', 'password', 'unit_number', 'switch_type']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'Missing required field: {field}'}), 400

        ip = data['ip']
        username = data['username']
        password = data['password']
        unit_number = int(data['unit_number'])
        switch_type = data['switch_type']

        logger.info(f"Extracting config from {ip}, unit={unit_number}, type={switch_type}")

        # Connect to switch
        with SSHManager(ip, username, password) as ssh:
            # Get interface brief
            brief_output = ssh.get_interface_brief()
            interfaces_list = ConfigParser.parse_interface_brief(brief_output)

            # Get detailed config for each interface
            interfaces_config = []

            # Determine max port based on switch type (24 or 48)
            max_port = int(switch_type)  # switch_type is "24" or "48"

            for iface in interfaces_list:
                # Use universal parser to check if it's a physical interface
                parsed = UniversalInterfaceParser.parse_interface_name(iface['name'])
                if parsed:  # If parseable, it's a physical interface
                    # Filter: Skip uplink ports (port > max_port)
                    # For 24-port: skip ports 25, 26, 27, 28
                    # For 48-port: skip ports 49, 50, 51, 52
                    if parsed['port'] > max_port:
                        logger.info(f"Skipping uplink port {iface['name']} (port {parsed['port']} > {max_port})")
                        continue

                    try:
                        config_output = ssh.get_interface_config(iface['name'])
                        config = ConfigParser.parse_interface_config(config_output)

                        # Translate interface name (accepts all port numbers, all vendors)
                        config = InterfaceTranslator.translate_full_config(
                            config, unit_number, switch_type
                        )

                        interfaces_config.append(config)
                    except Exception as e:
                        logger.warning(f"Failed to get config for {iface['name']}: {str(e)}")

        logger.info(f"Successfully extracted {len(interfaces_config)} interface configs via {ssh.connection_method}")

        return jsonify({
            'success': True,
            'interfaces': interfaces_config,
            'count': len(interfaces_config),
            'connection_method': ssh.connection_method
        })

    except Exception as e:
        logger.error(f"Config extraction failed: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/generate_config', methods=['POST'])
def generate_config():
    """Generate new switch configuration (simple version)"""
    try:
        data = request.get_json()

        # Validate input
        required_fields = ['interfaces', 'unit_number']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400

        interfaces = data['interfaces']
        unit_number = data['unit_number']

        logger.info(f"Generating config for {len(interfaces)} interfaces, unit={unit_number}")

        # Generate configuration
        generator = TemplateGenerator()
        config = generator.generate_config(interfaces, unit_number)

        filename = f"switch_unit_{unit_number}_config.txt"

        return jsonify({
            'success': True,
            'config': config,
            'filename': filename
        })

    except Exception as e:
        logger.error(f"Config generation failed: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/generate_config_complete', methods=['POST'])
def generate_config_complete():
    """Generate complete switch configuration with all VLANs and settings"""
    try:
        data = request.get_json()

        # Validate input
        required_fields = ['interfaces', 'switch_name', 'switch_ip', 'switch_gateway', 'admin_password']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400

        interfaces = data['interfaces']
        switch_name = data['switch_name']
        switch_ip = data['switch_ip']
        switch_gateway = data['switch_gateway']
        admin_password = data['admin_password']

        logger.info(f"Generating complete config for {len(interfaces)} interfaces")

        # Generate complete configuration
        generator = TemplateGenerator()
        config = generator.generate_complete_config(
            interfaces, switch_name, switch_ip, switch_gateway, admin_password
        )

        filename = f"{switch_name}_complete_config.txt"

        return jsonify({
            'success': True,
            'config': config,
            'filename': filename
        })

    except Exception as e:
        logger.error(f"Complete config generation failed: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/download_config', methods=['POST'])
def download_config():
    """Download generated configuration file"""
    try:
        data = request.get_json()

        if 'config' not in data or 'filename' not in data:
            return jsonify({'error': 'Missing config or filename'}), 400

        config_content = data['config']
        filename = data['filename']

        # Create in-memory file
        file_buffer = io.BytesIO()
        file_buffer.write(config_content.encode('utf-8'))
        file_buffer.seek(0)

        return send_file(
            file_buffer,
            mimetype='text/plain',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f"Config download failed: {str(e)}")
        return jsonify({'error': str(e)}), 500


@bp.route('/health')
def health():
    """Health check endpoint"""
    return jsonify({'status': 'ok'})


@bp.route('/api/process-stack', methods=['POST'])
def process_stack():
    """Process stack configuration - extract from multiple switches"""
    try:
        data = request.get_json()
        
        required_fields = ['stack_name', 'switches']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        stack_name = data['stack_name']
        switches = data['switches']
        
        if not switches or len(switches) == 0:
            return jsonify({'error': 'No switches provided'}), 400
        
        logger.info(f"Processing stack {stack_name} with {len(switches)} switches")
        
        all_interfaces = []
        stack_switches = []
        unique_vlans = set()
        
        for sw in switches:
            try:
                ip = sw['host']
                password = sw['password']
                unit = sw['unit']
                port_count = sw.get('port_count', 48)  # Default 48 porte se non specificato
                username = data.get('username', 'admin')

                logger.info(f"Extracting from {ip} unit {unit} (max {port_count} ports)")

                with SSHManager(ip, username, password) as ssh:
                    brief_output = ssh.get_interface_brief()
                    interfaces_list = ConfigParser.parse_interface_brief(brief_output)

                    switch_interfaces = []
                    for iface in interfaces_list:
                        parsed = UniversalInterfaceParser.parse_interface_name(iface['name'])
                        if parsed:
                            # Filter by port_count: skip ports > max (uplink ports)
                            if parsed['port'] > port_count:
                                logger.info(f"Skipping uplink port {iface['name']} (port {parsed['port']} > {port_count})")
                                continue

                            # For 8-port switches: GigabitEthernet is the uplink, skip it
                            if port_count == 8 and iface['name'].lower().startswith('gigabitethernet'):
                                logger.info(f"Skipping GigabitEthernet uplink {iface['name']} for 8-port switch")
                                continue

                            try:
                                config_output = ssh.get_interface_config(iface['name'])
                                config = ConfigParser.parse_interface_config(config_output)
                                config = InterfaceTranslator.translate_full_config(config, unit, str(port_count))
                                
                                switch_interfaces.append(config)
                                all_interfaces.append(config)
                                
                                if 'vlan' in config and config['vlan']:
                                    if isinstance(config['vlan'], list):
                                        unique_vlans.update(config['vlan'])
                                    else:
                                        unique_vlans.add(config['vlan'])
                                        
                            except Exception as e:
                                logger.warning(f"Failed config for {iface['name']}: {e}")
                
                stack_switches.append({
                    'stack_unit': unit,
                    'switch_ip': ip,
                    'ip_address': ip,
                    'port_count': port_count,
                    'interface_count': len(switch_interfaces),
                    'vlans': list(unique_vlans),
                    'connection_method': ssh.connection_method,
                    'error': None
                })
                
            except Exception as e:
                logger.error(f"Failed to process switch {sw.get('host', 'unknown')}: {e}")
                stack_switches.append({
                    'stack_unit': sw.get('unit', 0),
                    'switch_ip': sw.get('host', 'unknown'),
                    'ip_address': sw.get('host', 'unknown'),
                    'port_count': sw.get('port_count', 48),
                    'interface_count': 0,
                    'vlans': [],
                    'connection_method': None,
                    'error': str(e)
                })
        
        stack_summary = {
            'stack_name': stack_name,
            'switch_count': len(switches),
            'total_interfaces': len(all_interfaces),
            'vlans': sorted(list(unique_vlans)),
            'switches': stack_switches
        }
        
        # Store in session or temp file for step 3
        os.makedirs('/tmp/huawei_stack', exist_ok=True)
        import json
        with open(f'/tmp/huawei_stack/{stack_name}.json', 'w') as f:
            json.dump({'interfaces': all_interfaces, 'summary': stack_summary}, f)
        
        logger.info(f"Stack processing complete: {len(all_interfaces)} interfaces from {len(switches)} switches")
        
        return jsonify({
            'success': True,
            'message': f'Elaborati {len(switches)} switch con {len(all_interfaces)} interfacce',
            'stack_summary': stack_summary
        })
        
    except Exception as e:
        logger.error(f"Stack processing failed: {e}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/generate-stack-config', methods=['POST'])
def generate_stack_config():
    """Generate final stack configuration"""
    try:
        data = request.get_json()

        required_fields = ['stack_name', 'new_stack_ip', 'gateway', 'admin_password']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400

        stack_name = data['stack_name']
        new_stack_ip = data['new_stack_ip']
        gateway = data['gateway']
        admin_password = data['admin_password']
        lacp_enabled = data.get('lacp_enabled', False)
        stack_units = data.get('stack_units', [])

        # Load interfaces from temp file
        import json
        stack_file = f'/tmp/huawei_stack/{stack_name}.json'

        if not os.path.exists(stack_file):
            return jsonify({'error': 'Stack data not found. Please process stack first (Step 2)'}), 400

        with open(stack_file, 'r') as f:
            stack_data = json.load(f)

        interfaces = stack_data['interfaces']
        switches_info = stack_data.get('summary', {}).get('switches', [])

        logger.info(f"Generating stack config for {stack_name}: {len(interfaces)} interfaces, LACP={lacp_enabled}")

        generator = TemplateGenerator()
        config = generator.generate_complete_config(
            interfaces, stack_name, new_stack_ip, gateway, admin_password,
            lacp_enabled=lacp_enabled, stack_units=stack_units, switches_info=switches_info
        )
        
        filename = f"{stack_name}_config.txt"
        
        # Save to file
        config_path = f'/tmp/huawei_stack/{filename}'
        with open(config_path, 'w') as f:
            f.write(config)
        
        return jsonify({
            'success': True,
            'message': 'Configurazione stack generata con successo',
            'config_filename': filename,
            'config_size': len(config),
            'config_content': config
        })
        
    except Exception as e:
        logger.error(f"Stack config generation failed: {e}")
        return jsonify({'error': str(e)}), 500


@bp.route('/layer3')
def layer3():
    """Layer 3 migration page"""
    return render_template('layer3.html')


@bp.route('/api/parse-layer3-excel', methods=['POST'])
def parse_layer3_excel():
    """Parse Excel file for Layer 3 migration"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nessun file caricato'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nessun file selezionato'}), 400

        # Read Excel with openpyxl
        wb = load_workbook(file)
        ws = wb.active

        mappings = []
        eth_trunks = set()

        # Parse rows - skip first 2 rows (headers)
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            # Columns: 0=empty, 1=old_port, 2=new_port, 3=description, 4=eth_trunk, 5=link_type, 6=config1, 7=config2
            old_port = str(row[1]).strip() if row[1] else None
            new_port = str(row[2]).strip() if row[2] else None
            description = str(row[3]).strip() if row[3] else None
            eth_trunk = str(row[4]).strip() if row[4] else None
            link_type = str(row[5]).strip() if row[5] else None
            config1 = str(row[6]).strip() if row[6] else None
            config2 = str(row[7]).strip() if len(row) > 7 and row[7] else None

            # Skip if no new port mapping
            if not new_port or new_port == 'nan' or not new_port.startswith('interface'):
                continue

            # Extract port name from "interface XGigabitEthernet1/0/1"
            new_port_name = new_port.replace('interface ', '').strip()

            # Parse eth-trunk number
            eth_trunk_num = None
            if eth_trunk and eth_trunk != 'nan':
                match = re.search(r'eth-trunk\s*(\d+)', eth_trunk, re.IGNORECASE)
                if match:
                    eth_trunk_num = int(match.group(1))
                    eth_trunks.add(eth_trunk_num)

            # Parse link type
            port_link_type = None
            if link_type and 'trunk' in link_type.lower():
                port_link_type = 'trunk'
            elif link_type and 'access' in link_type.lower():
                port_link_type = 'access'

            # Parse VLAN from config
            vlan = None
            if config1:
                vlan_match = re.search(r'vlan\s+(\d+)', config1)
                if vlan_match:
                    vlan = int(vlan_match.group(1))

            # Check for pvid
            pvid = None
            if config1 and 'pvid' in config1.lower():
                pvid_match = re.search(r'pvid\s+vlan\s+(\d+)', config1)
                if pvid_match:
                    pvid = int(pvid_match.group(1))

            # Check for specific trunk vlans (not 2 to 4094)
            trunk_vlans = None
            if config2 and 'allow-pass' in config2.lower():
                if '2 to 4094' not in config2:
                    # Extract specific vlans
                    vlans_match = re.findall(r'(\d+)', config2)
                    if vlans_match:
                        trunk_vlans = [int(v) for v in vlans_match]

            mapping = {
                'old_port': old_port,
                'new_port': new_port_name,
                'description': description if description != 'nan' else None,
                'eth_trunk': eth_trunk_num,
                'link_type': port_link_type,
                'vlan': vlan,
                'pvid': pvid,
                'trunk_vlans': trunk_vlans
            }
            mappings.append(mapping)

        logger.info(f"Parsed {len(mappings)} port mappings, {len(eth_trunks)} eth-trunks")

        return jsonify({
            'success': True,
            'mappings': mappings,
            'eth_trunks': sorted(list(eth_trunks)),
            'total_mappings': len(mappings)
        })

    except Exception as e:
        logger.error(f"Excel parsing failed: {e}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/generate-layer3-config', methods=['POST'])
def generate_layer3_config():
    """Generate S6730 configuration from Excel mappings"""
    try:
        data = request.get_json()

        mappings = data.get('mappings', [])
        eth_trunks = data.get('eth_trunks', [])
        switch_name = data.get('switch_name', 'NEW_S6730')
        switch_ip = data.get('switch_ip', '')
        gateway = data.get('gateway', '')
        admin_password = data.get('admin_password', '')

        logger.info(f"Generating Layer3 config: {len(mappings)} interfaces, {len(eth_trunks)} eth-trunks")

        config_lines = []

        # Header
        config_lines.append("# Configurazione S6730-H48X6C")
        config_lines.append(f"# Switch: {switch_name}")
        config_lines.append(f"# Generata automaticamente")
        config_lines.append("#")
        config_lines.append("")
        config_lines.append("system-view")
        config_lines.append("")

        # Sysname
        config_lines.append(f"sysname {switch_name}")
        config_lines.append("")

        # Management VLAN and IP
        if switch_ip:
            config_lines.append("vlan 1000")
            config_lines.append(" description MGM_Switch")
            config_lines.append("#")
            config_lines.append("interface Vlanif1000")
            config_lines.append(f" ip address {switch_ip} 255.255.255.0")
            config_lines.append("#")

        if gateway:
            config_lines.append(f"ip route-static 0.0.0.0 0.0.0.0 {gateway}")
            config_lines.append("")

        # Admin user
        if admin_password:
            config_lines.append("aaa")
            config_lines.append(f"local-user admin password irreversible-cipher {admin_password}")
            config_lines.append("local-user admin privilege level 3")
            config_lines.append("local-user admin service-type terminal ssh")
            config_lines.append("q")
            config_lines.append("")

        # Eth-Trunks first
        if eth_trunks:
            config_lines.append("# Eth-Trunk Configuration")
            for trunk_num in sorted(eth_trunks):
                config_lines.append("#")
                config_lines.append(f"interface Eth-Trunk{trunk_num}")
                config_lines.append(" port link-type trunk")
                config_lines.append(" port trunk allow-pass vlan 2 to 4094")
                config_lines.append(" mode lacp")
                config_lines.append(" q")
            config_lines.append("")

        # Interface configurations
        config_lines.append("# Interface Configuration")

        stats = {'interfaces': 0, 'trunk_ports': 0, 'access_ports': 0, 'eth_trunks': len(eth_trunks)}

        for m in mappings:
            config_lines.append("#")
            config_lines.append(f"interface {m['new_port']}")

            if m.get('description'):
                desc = m['description']
                if not desc.startswith('description'):
                    desc = f"description {desc}"
                config_lines.append(f" {desc}")

            if m.get('eth_trunk'):
                config_lines.append(f" eth-trunk {m['eth_trunk']}")
            else:
                if m.get('link_type') == 'trunk':
                    config_lines.append(" port link-type trunk")
                    if m.get('pvid'):
                        config_lines.append(f" port trunk pvid vlan {m['pvid']}")
                    if m.get('trunk_vlans'):
                        vlans_str = ' '.join(str(v) for v in m['trunk_vlans'])
                        config_lines.append(f" port trunk allow-pass vlan {vlans_str}")
                    else:
                        config_lines.append(" port trunk allow-pass vlan 2 to 4094")
                    stats['trunk_ports'] += 1
                elif m.get('link_type') == 'access':
                    config_lines.append(" port link-type access")
                    if m.get('vlan'):
                        config_lines.append(f" port default vlan {m['vlan']}")
                    stats['access_ports'] += 1

            config_lines.append(" q")
            stats['interfaces'] += 1

        # Footer
        config_lines.append("#")
        config_lines.append("return")
        config_lines.append("")

        config = '\n'.join(config_lines)
        filename = f"{switch_name}_layer3_config.txt"

        return jsonify({
            'success': True,
            'config': config,
            'filename': filename,
            'stats': stats
        })

    except Exception as e:
        logger.error(f"Layer3 config generation failed: {e}")
        return jsonify({'error': str(e)}), 500
